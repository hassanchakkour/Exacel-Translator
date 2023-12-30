from translate import Translator
from openpyxl import workbook, load_workbook
from deep_translator import GoogleTranslator
import time

start_time = time.time()

wb = load_workbook('testQ.xlsx')

ws = wb.active

column_to_replace = 'E'

column_to_add = "F"

cell_count = 0

translator = GoogleTranslator(source='auto', target='en')

for row_number in range(3290, 3295):
    cell = ws[column_to_replace + str(row_number)]
    cellToAddFaculty = ws[column_to_add + str(row_number)]
    cell.value = translator.translate(cell.value)
    print(cell.value)
    splittedValue = cell.value.split(' - ', 3)
    print(len(splittedValue))
    if len(splittedValue) == 3:
        cell.value, cellToAddFaculty.value, thirdpart = splittedValue
    else:
        cell.value,cellToAddFaculty.value = cell.value.split(" - ")

    # cell_count += 1
    print(str(row_number))

    # print(f"Department : {cell.value}")
    # print(f"Faculty {cellToAddFaculty.value}")

elapsed_time = time.time() - start_time

print(f"Elapsed Time: {elapsed_time} seconds")
wb.save('testQ.xlsx')


